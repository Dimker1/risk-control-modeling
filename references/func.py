"""
数据处理基础函数模块
包含：数据加载、机构分析、缺失率计算、IV计算、PSI计算、报告导出
"""

import pandas as pd
import numpy as np
from typing import List, Dict, Tuple, Optional
import tqdm
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ============================================================
# 数据加载与格式化
# ============================================================

def get_dataset(data_path: str, date_col: str, y_col: str,
                org_col: str = None, key_cols: List[str] = None,
                drop_cols: List[str] = None,
                miss_vals: List[int] = None) -> pd.DataFrame:
    """加载并格式化数据

    Args:
        data_path: 数据文件路径（支持 parquet/csv/xlsx/pkl）
        date_col: 日期列名
        y_col: 标签列名
        org_col: 机构列名（可选，支持多机构分析）
        key_cols: 主键列名列表（用于去重）
        drop_cols: 需要删除的列名
        miss_vals: 异常值列表，替换为NaN，默认 [-1, -999, -1111]
    """
    if drop_cols is None:
        drop_cols = []
    if miss_vals is None:
        miss_vals = [-1, -999, -1111]
    if key_cols is None:
        key_cols = []

    # 多格式读取
    data = None
    for fmt, reader in [('parquet', pd.read_parquet), ('csv', pd.read_csv),
                         ('xlsx', pd.read_excel), ('pkl', pd.read_pickle)]:
        try:
            data = reader(data_path)
            print(f"  Data loaded via {fmt}: {data.shape}")
            break
        except Exception:
            continue

    if data is None:
        raise ValueError(f"Failed to load data from {data_path}")

    # 替换异常值为NaN
    data.replace({v: np.nan for v in miss_vals}, inplace=True)

    # 过滤有效标签
    data = data[data[y_col].isin([0, 1])]

    # 删除指定列
    data = data.drop(columns=[c for c in drop_cols if c in data.columns], errors='ignore')

    # 去重后删除主键列（主键不参与建模）
    if key_cols:
        valid_keys = [c for c in key_cols if c in data.columns]
        if valid_keys:
            data = data.drop_duplicates(subset=valid_keys)
            data = data.drop(columns=valid_keys, errors='ignore')

    # 删除常数特征
    const_cols = [c for c in data.columns if data[c].nunique() <= 1 and c != y_col]
    data = data.drop(columns=const_cols, errors='ignore')

    # 标准化列名
    rename_map = {date_col: 'new_date', y_col: 'new_target'}
    if org_col and org_col in data.columns:
        rename_map[org_col] = 'new_org'
    data.rename(columns=rename_map, inplace=True)

    # 日期格式统一为 YYYYMMDD
    data['new_date'] = data['new_date'].astype(str).str.replace('-', '', regex=False).str[:8]
    data['new_date_ym'] = data['new_date'].str[:6]

    # 如果没有机构列，统一设为 'default'
    if 'new_org' not in data.columns:
        data['new_org'] = 'default'

    print(f"  After preprocessing: {data.shape}, bad_rate={data['new_target'].mean():.4f}")
    return data


# ============================================================
# 机构样本分析
# ============================================================

def org_analysis(data: pd.DataFrame, oos_orgs: List[str] = None) -> pd.DataFrame:
    """机构样本统计分析

    Returns:
        每个机构每月的样本数、坏样本数、坏样本率统计表
    """
    if oos_orgs is None:
        oos_orgs = []

    stat = data.groupby(['new_org', 'new_date_ym']).agg(
        单月坏样本数=('new_target', 'sum'),
        单月总样本数=('new_target', 'count'),
        单月坏样率=('new_target', 'mean')
    ).reset_index()

    # 累计统计
    stat['总坏样本数'] = stat.groupby('new_org')['单月坏样本数'].transform('sum')
    stat['总样本数'] = stat.groupby('new_org')['单月总样本数'].transform('sum')
    stat['总坏样率'] = stat['总坏样本数'] / stat['总样本数']

    # 标记OOS机构
    stat['样本类型'] = stat['new_org'].apply(lambda x: '贷外' if x in oos_orgs else '建模')

    stat = stat.rename(columns={'new_org': '机构', 'new_date_ym': '年月'})
    stat = stat.sort_values(['样本类型', '机构', '年月']).reset_index(drop=True)

    return stat[['机构', '年月', '单月坏样本数', '单月总样本数', '单月坏样率',
                  '总坏样本数', '总样本数', '总坏样率', '样本类型']]


# ============================================================
# 缺失率计算
# ============================================================

def missing_check(data: pd.DataFrame, miss_vals: List[int] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """计算缺失率 - 包含整体和机构维度

    Args:
        data: 数据（需含 new_org 列）
        miss_vals: 额外视为缺失的值

    Returns:
        miss_detail: 缺失率明细（变量, 整体, org1, org2, ...）
        miss_overall: 整体缺失率汇总
    """
    if miss_vals is None:
        miss_vals = [-1, -999, -1111]

    exclude_cols = ['new_date', 'new_date_ym', 'new_target', 'new_org', 'record_id', 'target']
    cols = [c for c in data.columns if c not in exclude_cols]

    orgs = sorted(data['new_org'].unique())

    # 计算整体缺失率
    miss_overall_list = []
    for col in tqdm.tqdm(cols, desc="Missing rate"):
        rate = ((data[col].isin(miss_vals)) | (data[col].isna())).mean()
        miss_overall_list.append({'变量': col, '整体缺失率': round(rate, 4)})

    miss_overall = pd.DataFrame(miss_overall_list)

    # 计算机构维度缺失率，转宽表
    miss_detail_dict = {'变量': [], '整体': []}
    for org in orgs:
        miss_detail_dict[org] = []

    for col in tqdm.tqdm(cols, desc="Missing rate (by org)"):
        miss_detail_dict['变量'].append(col)
        overall_rate = ((data[col].isin(miss_vals)) | (data[col].isna())).mean()
        miss_detail_dict['整体'].append(round(overall_rate, 4))

        for org in orgs:
            org_data = data[data['new_org'] == org]
            rate = ((org_data[col].isin(miss_vals)) | (org_data[col].isna())).mean()
            miss_detail_dict[org].append(round(rate, 4))

    miss_detail = pd.DataFrame(miss_detail_dict)
    miss_detail = miss_detail.sort_values('整体', ascending=False).reset_index(drop=True)

    return miss_detail, miss_overall


# ============================================================
# IV 计算（支持多进程）
# ============================================================

def calculate_iv(data: pd.DataFrame, features: List[str], target: str = 'new_target',
                 n_jobs: int = 1, n_bins: int = 5) -> pd.DataFrame:
    """计算IV值 - 使用决策树分箱

    Args:
        data: 数据
        features: 特征列表
        target: 目标列名
        n_jobs: 并行进程数
        n_bins: 分箱数
    """
    from sklearn.tree import DecisionTreeClassifier

    def _calc_iv_single(f):
        try:
            X = data[[f]].copy()
            y = data[target].copy()

            # 处理缺失值
            mask = X[f].notna()
            X_valid = X[mask]
            y_valid = y[mask]

            if len(X_valid) < 50 or y_valid.nunique() < 2:
                return None

            # 决策树分箱
            dt = DecisionTreeClassifier(max_leaf_nodes=n_bins, min_samples_leaf=0.05, random_state=42)
            dt.fit(X_valid, y_valid)
            X_valid['bin'] = dt.apply(X_valid)

            # 计算WOE和IV
            iv_value = 0.0
            total_good = (y_valid == 0).sum()
            total_bad = (y_valid == 1).sum()

            if total_good == 0 or total_bad == 0:
                return None

            for bin_id in X_valid['bin'].unique():
                bin_mask = X_valid['bin'] == bin_id
                bin_good = ((y_valid == 0) & bin_mask).sum()
                bin_bad = ((y_valid == 1) & bin_mask).sum()

                good_pct = max(bin_good / total_good, 1e-6)
                bad_pct = max(bin_bad / total_bad, 1e-6)

                woe = np.log(good_pct / bad_pct)
                woe = np.clip(woe, -5, 5)  # 防溢出
                iv_value += (good_pct - bad_pct) * woe

            return {'变量': f, 'IV': round(iv_value, 4)}
        except Exception as e:
            return None

    if n_jobs > 1:
        from joblib import Parallel, delayed
        results = Parallel(n_jobs=n_jobs, verbose=0)(
            delayed(_calc_iv_single)(f) for f in features
        )
    else:
        results = [_calc_iv_single(f) for f in tqdm.tqdm(features, desc="IV calculation")]

    iv_list = [r for r in results if r is not None]

    if not iv_list:
        return pd.DataFrame(columns=['变量', 'IV'])

    return pd.DataFrame(iv_list).sort_values('IV', ascending=False).reset_index(drop=True)


def calculate_iv_by_org(data: pd.DataFrame, features: List[str],
                        target: str = 'new_target', n_jobs: int = 1) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """按机构计算IV，返回明细和整体IV

    Returns:
        iv_detail: IV明细（变量, 整体, org1, org2, ...）
        iv_overall: 整体IV表
    """
    orgs = sorted(data['new_org'].unique())

    # 整体IV
    iv_overall = calculate_iv(data, features, target, n_jobs=n_jobs)
    if len(iv_overall) == 0:
        return pd.DataFrame(columns=['变量', '整体']), iv_overall

    iv_overall = iv_overall.rename(columns={'IV': 'IV值'})

    # 机构IV
    iv_by_org = {}
    for org in orgs:
        org_data = data[data['new_org'] == org]
        org_iv = calculate_iv(org_data, features, target, n_jobs=1)
        if len(org_iv) > 0:
            iv_by_org[org] = dict(zip(org_iv['变量'], org_iv['IV']))

    # 转宽表
    iv_detail_dict = {'变量': [], '整体': []}
    for org in orgs:
        iv_detail_dict[org] = []

    all_vars = iv_overall['变量'].tolist()
    for var in all_vars:
        iv_detail_dict['变量'].append(var)
        var_overall = iv_overall[iv_overall['变量'] == var]['IV值'].values
        iv_detail_dict['整体'].append(var_overall[0] if len(var_overall) > 0 else None)

        for org in orgs:
            iv_detail_dict[org].append(iv_by_org.get(org, {}).get(var, None))

    iv_detail = pd.DataFrame(iv_detail_dict)
    iv_detail = iv_detail.sort_values('整体', ascending=False).reset_index(drop=True)

    return iv_detail, iv_overall


# ============================================================
# PSI 计算
# ============================================================

def calculate_psi(expected, actual, bins=10):
    """计算PSI (Population Stability Index)

    Args:
        expected: 基准期数据
        actual: 比较期数据
        bins: 分箱数
    """
    expected = np.array(expected, dtype=float)
    actual = np.array(actual, dtype=float)

    # 去除NaN
    expected = expected[~np.isnan(expected)]
    actual = actual[~np.isnan(actual)]

    if len(expected) == 0 or len(actual) == 0:
        return 0.0

    # 基于expected分箱
    breakpoints = np.arange(0, bins + 1) / bins * 100
    breakpoints = np.percentile(expected, breakpoints)
    breakpoints[0] = -np.inf
    breakpoints[-1] = np.inf

    expected_bins = np.histogram(expected, bins=breakpoints)[0]
    actual_bins = np.histogram(actual, bins=breakpoints)[0]

    expected_pct = expected_bins / len(expected)
    actual_pct = actual_bins / len(actual)

    # 避免除0
    expected_pct = np.clip(expected_pct, 1e-6, None)
    actual_pct = np.clip(actual_pct, 1e-6, None)

    psi = np.sum((actual_pct - expected_pct) * np.log(actual_pct / expected_pct))
    return round(psi, 4)


def calculate_psi_by_org(data: pd.DataFrame, features: List[str],
                         psi_threshold: float = 0.1,
                         max_months_ratio: float = 1/3,
                         max_orgs: int = 6,
                         min_sample: int = 100,
                         n_jobs: int = 1) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """按机构按月计算PSI

    Returns:
        data: 处理后的数据
        psi_detail: PSI明细（机构, 变量, 月份, PSI值）
        psi_process: PSI处理表（需剔除的特征）
    """
    orgs = sorted(data['new_org'].unique())
    psi_detail_list = []
    psi_summary = []

    for org in orgs:
        org_data = data[data['new_org'] == org]
        months = sorted(org_data['new_date_ym'].unique())

        if len(months) < 2:
            continue

        # 第一个月作为基准
        base_month = months[0]
        base_data = org_data[org_data['new_date_ym'] == base_month]

        for f in features:
            if f not in org_data.columns:
                continue

            # 初始月份PSI为0
            psi_detail_list.append({
                '机构': org, '变量': f, '月份': base_month, 'PSI值': 0.0
            })

            unstable_months = 0
            total_months = 0

            for month in months[1:]:
                month_data = org_data[org_data['new_date_ym'] == month]
                if len(base_data) < min_sample or len(month_data) < min_sample:
                    psi_detail_list.append({
                        '机构': org, '变量': f, '月份': month, 'PSI值': None
                    })
                    continue

                psi_val = calculate_psi(base_data[f].values, month_data[f].values)
                psi_detail_list.append({
                    '机构': org, '变量': f, '月份': month, 'PSI值': psi_val
                })

                total_months += 1
                if psi_val > psi_threshold:
                    unstable_months += 1

            if total_months > 0:
                unstable_threshold = max(1, int(total_months * max_months_ratio))
                is_unstable = unstable_months >= unstable_threshold
                psi_summary.append({
                    '机构': org, '变量': f, '不稳定月份数': unstable_months,
                    '总月份数': total_months, '是否不稳定': is_unstable
                })

    psi_detail = pd.DataFrame(psi_detail_list)

    if not psi_summary:
        return data, psi_detail, pd.DataFrame(columns=['变量', '处理原因', '不稳定机构'])

    psi_summary_df = pd.DataFrame(psi_summary)

    # 统计每个变量在多少个机构不稳定
    unstable_orgs_dict = {}
    for var in psi_summary_df['变量'].unique():
        var_data = psi_summary_df[psi_summary_df['变量'] == var]
        unstable_orgs = var_data[var_data['是否不稳定'] == True]['机构'].tolist()
        unstable_orgs_dict[var] = unstable_orgs

    org_count = len(orgs)
    channel_summary = psi_summary_df.groupby('变量').agg(
        不稳定机构数=('是否不稳定', 'sum')
    ).reset_index()
    channel_summary['需处理'] = channel_summary['不稳定机构数'] >= max_orgs
    channel_summary['处理原因'] = channel_summary.apply(
        lambda x: f'PSI不稳定机构数{x["不稳定机构数"]}/{org_count}' if x['需处理'] else '', axis=1
    )
    channel_summary['不稳定机构'] = channel_summary['变量'].apply(
        lambda x: ','.join(unstable_orgs_dict.get(x, []))
    )

    psi_process = channel_summary[channel_summary['需处理']].copy()
    psi_process = psi_process[['变量', '处理原因', '不稳定机构']].reset_index(drop=True)

    # 剔除不稳定特征
    if len(psi_process) > 0:
        to_drop = psi_process['变量'].tolist()
        data = data.drop(columns=[c for c in to_drop if c in data.columns], errors='ignore')

    return data, psi_detail, psi_process


# ============================================================
# IV/PSI 分布统计
# ============================================================

def iv_distribution_by_org(iv_detail: pd.DataFrame, oos_orgs: list = None,
                           iv_bins: list = None) -> pd.DataFrame:
    """按机构统计IV分布

    Args:
        iv_detail: IV明细表（含变量, 整体, org1, org2, ...列）
        iv_bins: IV区间边界，默认 [0, 0.02, 0.05, 0.1, inf]
    """
    if oos_orgs is None:
        oos_orgs = []
    if iv_bins is None:
        iv_bins = [0, 0.02, 0.05, 0.1, float('inf')]

    org_cols = [c for c in iv_detail.columns if c not in ['变量', '整体']]
    bin_labels = [f'[{iv_bins[i]}, {iv_bins[i+1]})' for i in range(len(iv_bins) - 1)]

    result = []
    for org in org_cols:
        org_iv = iv_detail[org].dropna()
        total_vars = len(org_iv)
        org_type = '贷外' if org in oos_orgs else '建模'

        for i in range(len(iv_bins) - 1):
            lower, upper = iv_bins[i], iv_bins[i + 1]
            if upper == float('inf'):
                count = (org_iv >= lower).sum()
            else:
                count = ((org_iv >= lower) & (org_iv < upper)).sum()
            ratio = count / total_vars if total_vars > 0 else 0
            result.append({
                '机构': org, '类型': org_type, 'IV区间': bin_labels[i],
                '变量个数': count, '占比': f'{ratio:.2%}'
            })

    return pd.DataFrame(result)


def psi_distribution_by_org(psi_detail: pd.DataFrame, oos_orgs: list = None,
                            psi_bins: list = None) -> pd.DataFrame:
    """按机构统计PSI分布"""
    if oos_orgs is None:
        oos_orgs = []
    if psi_bins is None:
        psi_bins = [0, 0.05, 0.1, float('inf')]

    bin_labels = [f'[{psi_bins[i]}, {psi_bins[i+1]})' for i in range(len(psi_bins) - 1)]
    result = []

    for org in psi_detail['机构'].unique():
        org_data = psi_detail[psi_detail['机构'] == org]
        org_type = '贷外' if org in oos_orgs else '建模'
        var_max_psi = org_data.groupby('变量')['PSI值'].max().dropna()
        total_vars = len(var_max_psi)

        for i in range(len(psi_bins) - 1):
            lower, upper = psi_bins[i], psi_bins[i + 1]
            if upper == float('inf'):
                count = (var_max_psi >= lower).sum()
            else:
                count = ((var_max_psi >= lower) & (var_max_psi < upper)).sum()
            ratio = count / total_vars if total_vars > 0 else 0
            result.append({
                '机构': org, '类型': org_type, 'PSI区间': bin_labels[i],
                '变量个数': count, '占比': f'{ratio:.2%}'
            })

    return pd.DataFrame(result)


def value_ratio_distribution_by_org(data: pd.DataFrame, features: List[str],
                                    oos_orgs: list = None,
                                    value_bins: list = None) -> pd.DataFrame:
    """按机构统计有值率分布"""
    if oos_orgs is None:
        oos_orgs = []
    if value_bins is None:
        value_bins = [0, 0.15, 0.35, 0.65, 0.95, 1.0]

    bin_labels = ['[0, 15%)', '[15%, 35%)', '[35%, 65%)', '[65%, 95%)', '[95%, 100%]']
    result = []

    for org in data['new_org'].unique():
        org_data = data[data['new_org'] == org]
        org_type = '贷外' if org in oos_orgs else '建模'

        value_ratios = {}
        for f in features:
            if f in org_data.columns:
                non_null_count = org_data[f].notna().sum()
                total_count = len(org_data)
                value_ratios[f] = non_null_count / total_count if total_count > 0 else 0

        total_vars = len(value_ratios)
        for i in range(len(value_bins) - 1):
            lower, upper = value_bins[i], value_bins[i + 1]
            if upper == 1.0:
                count = sum(1 for v in value_ratios.values() if lower <= v <= upper)
            else:
                count = sum(1 for v in value_ratios.values() if lower <= v < upper)
            ratio = count / total_vars if total_vars > 0 else 0
            result.append({
                '机构': org, '类型': org_type, '有值率区间': bin_labels[i],
                '变量个数': count, '占比': f'{ratio:.2%}'
            })

    return pd.DataFrame(result)


# ============================================================
# Excel 报告导出
# ============================================================

def export_report_xlsx(filepath: str, steps: list, params: dict = None):
    """导出Excel清洗报告 - 每步一个sheet

    Args:
        filepath: 输出路径
        steps: 清洗步骤列表 [(步骤名, DataFrame), ...]
        params: 超参数字典
    """
    if not HAS_OPENPYXL:
        print("  [WARN] openpyxl not installed, skipping Excel report")
        return

    from openpyxl import load_workbook

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    try:
        wb = load_workbook(filepath)
    except Exception:
        wb = Workbook()
        wb.remove(wb.active)

    if params is None:
        params = {}

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # 汇总sheet
    if '汇总' in wb.sheetnames:
        del wb['汇总']
    ws = wb.create_sheet('汇总', 0)
    ws['A1'] = '数据清洗报告'
    ws['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'] = '说明: 每步独立执行，数据不删除，仅统计不满足条件的特征'
    ws['A4'] = '步骤'
    ws['B4'] = '操作详情'
    ws['C4'] = '操作结果'
    ws['D4'] = '条件'

    row = 5
    filter_steps = ['Step4-异常月份处理', 'Step6-高缺失率处理', 'Step7-IV处理',
                    'Step8-PSI处理', 'Step9-Null Importance处理', 'Step10-高相关性剔除']

    for name, df in steps:
        display_name = name
        for prefix in ['Step4-', 'Step6-', 'Step7-', 'Step8-', 'Step9-', 'Step10-']:
            display_name = display_name.replace(prefix, '')

        if name in filter_steps and df is not None and len(df) > 0:
            ws.cell(row, 1, name)
            ws.cell(row, 2, display_name)
            if '变量' in df.columns:
                ws.cell(row, 3, f'剔除 {len(df)} 个特征')
            else:
                ws.cell(row, 3, f'剔除 {len(df)} 条记录')
            # 条件说明
            condition = _get_step_condition(name, params)
            ws.cell(row, 4, condition)
            row += 1
        elif name not in ['Step7-IV明细', 'Step7-IV分布统计', 'Step8-PSI明细',
                          'Step8-PSI分布统计', 'Step5-有值率分布统计']:
            ws.cell(row, 1, name)
            ws.cell(row, 2, display_name)
            ws.cell(row, 3, f'{len(df)} 条记录' if df is not None and len(df) > 0 else '空')
            ws.cell(row, 4, '')
            row += 1

    # 各步骤详情sheet
    for name, df in steps:
        if df is None or len(df) == 0:
            continue
        sheet_name = name[:31]  # Excel sheet名最长31字符
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws_detail = wb.create_sheet(sheet_name)

        for j, col in enumerate(df.columns):
            ws_detail.cell(1, j + 1, col)
            ws_detail.cell(1, j + 1).fill = header_fill
            ws_detail.cell(1, j + 1).font = header_font

        for i, (_, row_data) in enumerate(df.iterrows()):
            for j, val in enumerate(row_data):
                ws_detail.cell(i + 2, j + 1, val if val is not None else '')

    wb.save(filepath)
    print(f"  Report saved: {filepath}")


def _get_step_condition(step_name: str, params: dict) -> str:
    """获取步骤的条件说明"""
    conditions = {
        'Step4-异常月份处理': f"坏样本数<{params.get('min_ym_bad_sample', 10)} 或 总样本数<{params.get('min_ym_sample', 500)} 的月份剔除",
        'Step6-高缺失率处理': f"整体缺失率>{params.get('missing_ratio', 0.6)} 的特征剔除",
        'Step7-IV处理': f"整体IV<{params.get('overall_iv_threshold', 0.1)} 或 {params.get('max_org_threshold', 2)}+个机构IV<{params.get('org_iv_threshold', 0.1)} 的特征剔除",
        'Step8-PSI处理': f"PSI>{params.get('psi_threshold', 0.1)} 不稳定月份占比>{params.get('max_months_ratio', 1/3):.0%} 的机构>{params.get('max_orgs', 6)}个则剔除",
        'Step9-Null Importance处理': f"原始gain与置换gain差值<{params.get('gain_threshold', 50)} 视为噪声剔除",
        'Step10-高相关性剔除': f"相关性>{params.get('max_corr', 0.8)} 的特征剔除，保留gain前{params.get('top_n_keep', 20)}名",
    }
    return conditions.get(step_name, '')


import os  # ensure os is imported for export_report_xlsx
